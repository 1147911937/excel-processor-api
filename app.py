from flask import Flask, request, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64
import os

app = Flask(__name__)

# === 你的模板内容（来自 Office Scripts）===
TEMPLATE_A1J2 = [
    ["Print date :  1/1/2025  11:03:03PM", "Full vs Partial Pallet Report", "FACILITY - 122", "Customer - COB1", "Item - ALL", "Pull Type - Detail", "", "", "", ""],
    ["ITEM", "LOT", "LOCATION", "LPN", "INVSTATUS", "STATUS", "FULL PLT QTY", "QUANTITY", "UOM", "FULL/PARTIAL"]
]

@app.route('/health')
def health():
    return jsonify({"status": "OK"})

@app.route('/process', methods=['POST'])
def process_excel():
    try:
        data = request.get_json()
        if not data or 'file' not in data:
            return jsonify({"error": "Missing 'file' (base64)"}), 400

        file_data = base64.b64decode(data['file'])
        
        # 读取 .xls
        import xlrd
        book = xlrd.open_workbook(file_contents=file_data)
        sheet = book.sheet_by_index(0)

        # 创建新工作簿
        wb = openpyxl.Workbook()
        ws = wb.active

        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                ws.cell(row=r+1, column=c+1, value=sheet.cell_value(r, c))

        a1_val = ws['A8'].value if ws['A8'].value is not None else ""

        # 清空 1-12 行
        for row in ws.iter_rows(min_row=1, max_row=12):
            for cell in row:
                cell.value = None

        # 删除 3-12 行（10行）
        for _ in range(10):
            ws.delete_rows(3)

        # 写入模板
        for i, row in enumerate(TEMPLATE_A1J2):
            for j, val in enumerate(row):
                ws.cell(row=i+1, column=j+1, value=val)

        # 格式化 A2:J2
        font = Font(name="Times New Roman", size=8, bold=True)
        for col in range(1, 11):
            cell = ws.cell(row=2, column=col)
            cell.font = font
            cell.alignment = Alignment(wrap_text=True)

        ws['A1'] = a1_val

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Processed_Report.xlsx'
        )

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": "Processing failed"}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 8000))
    app.run(host='0.0.0.0', port=port)